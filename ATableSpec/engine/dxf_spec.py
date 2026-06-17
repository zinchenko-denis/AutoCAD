#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dxf_spec — извлечение и агрегация данных из блоков AutoCAD (DXF)
без зависимости от надстройки СПДС.

Воспроизводит ключевую функциональность СПДС-таблиц:
  1. читает блоки (INSERT) и их атрибуты из чертежа;
  2. по внешнему конфигу проецирует сырые атрибуты в канонические поля
     (марка / артикул / длина / размер заполнения / параметры торцовки);
  3. группирует, суммирует и сортирует записи в ведомости и спецификации.

Конфиг (YAML) играет роль "шаблона СПДС": схема описана данными, не кодом,
поэтому под другой проект/набор блоков правится только конфиг, не модуль.

Зависимости: ezdxf (чтение DXF), PyYAML (конфиг), openpyxl (экспорт XLSX, опц.).

Автор-нейтральный, переносимый модуль. CLI:
    python dxf_spec.py CHERTEZH.dxf -c mapping.yaml -o out [--xlsx]
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import ezdxf
import yaml

import atspec_report  # ядро «своего отчёта» (выражения + шаблон), версионно-независимое


# --------------------------------------------------------------------------- #
#  Структуры данных
# --------------------------------------------------------------------------- #
@dataclass
class RawBlock:
    """Сырая вставка блока — то, что физически лежит в чертеже."""
    name: str                       # имя блока (часто анонимное, *U…)
    layer: str
    x: float
    y: float
    rotation: float
    xscale: float
    yscale: float
    attributes: Dict[str, str] = field(default_factory=dict)

    # --- контракт обмена с внешним источником (C#-плагин -> JSON -> движок) ---
    def to_record(self) -> dict:
        return {
            "name": self.name, "layer": self.layer,
            "x": self.x, "y": self.y, "rotation": self.rotation,
            "xscale": self.xscale, "yscale": self.yscale,
            "attributes": self.attributes,
        }

    @staticmethod
    def from_record(d: dict) -> "RawBlock":
        attrs = d.get("attributes") or {}
        return RawBlock(
            name=str(d.get("name", "")),
            layer=str(d.get("layer", "")),
            x=float(d.get("x", 0) or 0),
            y=float(d.get("y", 0) or 0),
            rotation=float(d.get("rotation", 0) or 0),
            xscale=float(d.get("xscale", 1) or 1),
            yscale=float(d.get("yscale", 1) or 1),
            attributes={str(k): ("" if v is None else str(v)) for k, v in attrs.items()},
        )


@dataclass
class Element:
    """Канонический элемент после маппинга по конфигу."""
    etype: str                      # тип (ригель/стойка/заполнение/…)
    layer: str
    block: str
    mark: Optional[str] = None      # марка/маркировка/обозначение
    article: Optional[str] = None   # артикул профиля
    length: Optional[float] = None  # длина, мм
    size: Optional[str] = None      # сырой размер заполнения "ШхВ"
    width: Optional[float] = None   # ширина из size, мм
    height: Optional[float] = None  # высота из size, мм
    extras: Dict[str, str] = field(default_factory=dict)  # прочие атрибуты (keep_attrs)
    raw: Dict[str, str] = field(default_factory=dict)     # ВСЕ исходные атрибуты блока
    x: float = 0.0
    y: float = 0.0

    # доступ к полю по имени: канон -> extras -> сырые атрибуты блока
    def get(self, name: str) -> Any:
        canon = {
            "type": self.etype, "etype": self.etype,
            "layer": self.layer, "block": self.block,
            "mark": self.mark, "article": self.article,
            "length": self.length, "size": self.size,
            "width": self.width, "height": self.height,
            "x": self.x, "y": self.y,
        }
        if name in canon:
            return canon[name]
        if name in self.extras:
            return self.extras[name]
        return self.raw.get(name)


@dataclass
class Spec:
    """Готовая ведомость: имя, заголовок, колонки и строки."""
    name: str
    title: str
    columns: List[str]
    rows: List[Dict[str, Any]]
    total_qty: int = 0
    total_length_mm: Optional[float] = None


# --------------------------------------------------------------------------- #
#  Парсеры значений
# --------------------------------------------------------------------------- #
def parse_number(text: Optional[str]) -> Optional[float]:
    """'2748.00' / '540,5' / '' -> float | None. Терпим к запятой и мусору."""
    if text is None:
        return None
    s = str(text).strip().replace(",", ".")
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None


def parse_size(text: Optional[str], separators: List[str]) -> Tuple[Optional[float], Optional[float]]:
    """'510Х1170' -> (510.0, 1170.0). Разделители из конфига (кириллич. 'Х' и пр.)."""
    if not text:
        return None, None
    s = str(text).strip()
    pattern = "[" + re.escape("".join(separators)) + "]"
    parts = [p for p in re.split(pattern, s) if p.strip()]
    if len(parts) >= 2:
        return parse_number(parts[0]), parse_number(parts[1])
    return parse_number(s), None


# --------------------------------------------------------------------------- #
#  1. Extractor — чтение блоков из DXF
# --------------------------------------------------------------------------- #
class BlockExtractor:
    """Читает все INSERT из пространства модели и отдаёт RawBlock."""

    def __init__(self, dxf_path: str):
        self.dxf_path = dxf_path
        self.doc = ezdxf.readfile(dxf_path)   # ezdxf сам декодирует ANSI_1251

    def extract(self) -> List[RawBlock]:
        msp = self.doc.modelspace()
        out: List[RawBlock] = []
        for ins in msp.query("INSERT"):
            attrs = {a.dxf.tag: (a.dxf.text or "") for a in ins.attribs}
            out.append(RawBlock(
                name=ins.dxf.name,
                layer=ins.dxf.layer,
                x=float(ins.dxf.insert.x),
                y=float(ins.dxf.insert.y),
                rotation=float(getattr(ins.dxf, "rotation", 0.0) or 0.0),
                xscale=float(getattr(ins.dxf, "xscale", 1.0) or 1.0),
                yscale=float(getattr(ins.dxf, "yscale", 1.0) or 1.0),
                attributes=attrs,
            ))
        return out


# --------------------------------------------------------------------------- #
#  2. Mapper — сырые атрибуты -> канонический Element (по конфигу)
# --------------------------------------------------------------------------- #
class ElementMapper:
    def __init__(self, config: dict):
        self.layers: Dict[str, dict] = config.get("element_layers", {})
        self.field_map: Dict[str, List[str]] = config.get("field_map", {})
        self.keep_attrs: List[str] = config.get("keep_attrs", [])
        self.size_seps: List[str] = config.get("size_separators", ["Х", "X", "x", "х", "*"])

    def _first_nonempty(self, attrs: Dict[str, str], tags: List[str]) -> Optional[str]:
        for t in tags:
            v = (attrs.get(t) or "").strip()
            if v:
                return v
        return None

    def map(self, raw: RawBlock) -> Optional[Element]:
        """None, если слой не объявлен элементным (рамка/логотип/служебное)."""
        rule = self.layers.get(raw.layer)
        if rule is None:
            return None
        a = raw.attributes

        mark = self._first_nonempty(a, self.field_map.get("mark", []))
        article = self._first_nonempty(a, self.field_map.get("article", []))
        length = parse_number(self._first_nonempty(a, self.field_map.get("length", [])))
        size = self._first_nonempty(a, self.field_map.get("size", []))
        width, height = parse_size(size, self.size_seps)

        extras = {k: a[k] for k in self.keep_attrs if k in a}

        return Element(
            etype=rule.get("type", raw.layer),
            layer=raw.layer, block=raw.name,
            mark=mark, article=article, length=length,
            size=size, width=width, height=height,
            extras=extras, raw=dict(a), x=raw.x, y=raw.y,
        )

    def map_all(self, blocks: List[RawBlock]) -> Tuple[List[Element], List[RawBlock]]:
        """Возвращает (элементы, пропущенные блоки на неэлементных слоях)."""
        elements, skipped = [], []
        for b in blocks:
            el = self.map(b)
            (elements if el is not None else skipped).append(el if el is not None else b)
        return elements, skipped


# --------------------------------------------------------------------------- #
#  3. SpecBuilder — группировка / суммирование / сортировка
# --------------------------------------------------------------------------- #
def _sort_key_value(v: Any):
    """Числа сортируем как числа, остальное — как строки; None в конец."""
    if v is None:
        return (2, "")
    if isinstance(v, (int, float)):
        return (0, v)
    return (1, str(v))


def _cmp(a: Any, op: str, b: Any) -> bool:
    """Сравнение значения a с эталоном b по оператору op.
    По умолчанию строковое (регистронезависимое); для >,<,>=,<= — числовое."""
    sa = "" if a is None else str(a).strip()
    sb = "" if b is None else str(b).strip()
    if op in ("=", "=="):
        return sa.lower() == sb.lower()
    if op in ("!=", "<>"):
        return sa.lower() != sb.lower()
    if op == "contains":
        return sb.lower() in sa.lower()
    if op in (">", "<", ">=", "<="):
        na, nb = parse_number(sa), parse_number(sb)
        if na is None or nb is None:
            return False
        return {">": na > nb, "<": na < nb, ">=": na >= nb, "<=": na <= nb}[op]
    return True


def _passes(e: "Element", include_types, include_layers, filters) -> bool:
    """Проходит ли элемент фильтры запроса (источник по типам/слоям + условия AND)."""
    if include_types and e.etype not in include_types:
        return False
    if include_layers and e.layer not in include_layers:
        return False
    for f in (filters or []):
        field = f.get("field")
        if not field:
            continue
        if not _cmp(e.get(field), f.get("op", "="), f.get("value", "")):
            return False
    return True


class SpecBuilder:
    def __init__(self, config: dict):
        self.reports: List[dict] = config.get("reports", [])

    def build(self, report_cfg: dict, elements: List[Element]) -> Spec:
        include = set(report_cfg.get("include_types", []))
        include_layers = set(report_cfg.get("include_layers", []))
        filters = report_cfg.get("filters", [])
        group_by: List[str] = report_cfg["group_by"]
        measures: Dict[str, str] = report_cfg.get("measures", {"qty": "count"})
        sort_by: List[str] = report_cfg.get("sort_by", group_by)

        subset = [e for e in elements if _passes(e, include, include_layers, filters)]

        # агрегируем
        buckets: "OrderedDict[tuple, dict]" = OrderedDict()
        for e in subset:
            key = tuple(e.get(f) for f in group_by)
            b = buckets.get(key)
            if b is None:
                b = {f: e.get(f) for f in group_by}
                b["_count"] = 0
                b["_sum_length"] = 0.0
                buckets[key] = b
            b["_count"] += 1
            if e.length is not None:
                b["_sum_length"] += e.length

        # формируем строки и считаем меры
        rows: List[Dict[str, Any]] = []
        for b in buckets.values():
            row = {f: b[f] for f in group_by}
            for mname, mtype in measures.items():
                if mtype == "count":
                    row[mname] = b["_count"]
                elif mtype == "sum_length":
                    row[mname] = round(b["_sum_length"], 2)
                else:
                    raise ValueError(f"неизвестная мера: {mtype}")
            rows.append(row)

        # сортировка
        rows.sort(key=lambda r: tuple(_sort_key_value(r.get(f)) for f in sort_by))

        columns = list(group_by) + list(measures.keys())
        total_qty = sum(b["_count"] for b in buckets.values())
        total_len = round(sum(b["_sum_length"] for b in buckets.values()), 2)
        has_len = any(m == "sum_length" for m in measures.values())

        return Spec(
            name=report_cfg["name"],
            title=report_cfg.get("title", report_cfg["name"]),
            columns=columns, rows=rows,
            total_qty=total_qty,
            total_length_mm=(total_len if has_len else None),
        )

    def build_all(self, elements: List[Element]) -> List[Spec]:
        return [self.build(rc, elements) for rc in self.reports]


# --------------------------------------------------------------------------- #
#  Экспорт
# --------------------------------------------------------------------------- #
def write_csv(spec: Spec, path: str) -> None:
    # utf-8-sig -> Excel корректно открывает кириллицу
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(spec.columns)
        for r in spec.rows:
            w.writerow(["" if r.get(c) is None else r.get(c) for c in spec.columns])
        w.writerow([])
        w.writerow(["ИТОГО позиций:", spec.total_qty])
        if spec.total_length_mm is not None:
            w.writerow(["Сумм. длина, мм:", spec.total_length_mm])
            w.writerow(["Сумм. длина, м:", round(spec.total_length_mm / 1000.0, 3)])


def write_raw_csv(elements: List[Element], path: str) -> None:
    cols = ["type", "layer", "block", "mark", "article",
            "length", "width", "height", "size", "x", "y"]
    extra_keys = sorted({k for e in elements for k in e.extras})
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(cols + extra_keys)
        for e in elements:
            base = [e.get(c) if e.get(c) is not None else "" for c in cols]
            w.writerow(base + [e.extras.get(k, "") for k in extra_keys])


def write_xlsx(specs: List[Spec], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for spec in specs:
        ws = wb.create_sheet(title=spec.name[:31])
        ws.append([spec.title])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append(spec.columns)
        hdr = ws.max_row
        for c in range(1, len(spec.columns) + 1):
            cell = ws.cell(hdr, c)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for r in spec.rows:
            ws.append(["" if r.get(c) is None else r.get(c) for c in spec.columns])
            for c in range(1, len(spec.columns) + 1):
                ws.cell(ws.max_row, c).border = border
        ws.append([])
        ws.append(["ИТОГО позиций:", spec.total_qty])
        if spec.total_length_mm is not None:
            ws.append(["Сумм. длина, мм:", spec.total_length_mm])
            ws.append(["Сумм. длина, м:", round(spec.total_length_mm / 1000.0, 3)])
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)
    wb.save(path)


# --------------------------------------------------------------------------- #
#  Печать в консоль
# --------------------------------------------------------------------------- #
def print_spec(spec: Spec) -> None:
    print(f"\n=== {spec.title}  [{spec.name}] ===")
    if not spec.rows:
        print("  (нет элементов)")
        return
    widths = {c: max(len(str(c)), max((len(str(r.get(c, ""))) for r in spec.rows), default=0))
              for c in spec.columns}
    line = "  " + " | ".join(str(c).ljust(widths[c]) for c in spec.columns)
    print(line)
    print("  " + "-" * (len(line) - 2))
    for r in spec.rows:
        print("  " + " | ".join(str(r.get(c, "") if r.get(c) is not None else "")
                                 .ljust(widths[c]) for c in spec.columns))
    tail = f"  ИТОГО позиций: {spec.total_qty}"
    if spec.total_length_mm is not None:
        tail += f" | сумм. длина: {spec.total_length_mm:.2f} мм = {spec.total_length_mm/1000:.3f} м"
    print(tail)


# --------------------------------------------------------------------------- #
#  Гибрид: обработка записей блоков, полученных извне (C#-плагин из живого DWG)
# --------------------------------------------------------------------------- #
def records_to_specs(records: List[dict], config: dict):
    """Список record-словарей -> (specs, elements, skipped). Логика та же,
    что и для DXF; меняется только источник блоков."""
    mapper = ElementMapper(config)
    blocks = [RawBlock.from_record(r) for r in records]
    elements, skipped = mapper.map_all(blocks)
    specs = SpecBuilder(config).build_all(elements)
    return specs, elements, skipped


def spec_to_dict(spec: Spec) -> dict:
    return {
        "name": spec.name, "title": spec.title,
        "columns": spec.columns, "rows": spec.rows,
        "total_qty": spec.total_qty, "total_length_mm": spec.total_length_mm,
    }


def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# поля, по которым осмысленно фильтровать/группировать (помимо сырых атрибутов)
_DERIVED_FIELDS = ["layer", "type", "width", "height", "length"]


def describe(elements: List[Element], config: dict) -> dict:
    """Метаданные для построителя запроса в плагине: какие типы/слои/поля/значения
    есть в выборке, и какие пресеты заданы в конфиге."""
    types, layers = [], []
    attr_tags = []
    values: Dict[str, set] = {}

    def add_val(field, v):
        if v is None or str(v).strip() == "":
            return
        values.setdefault(field, set()).add(str(v))

    for e in elements:
        if e.etype not in types:
            types.append(e.etype)
        if e.layer not in layers:
            layers.append(e.layer)
        add_val("type", e.etype)
        add_val("layer", e.layer)
        add_val("width", e.width)
        add_val("height", e.height)
        add_val("length", e.length)
        for tag, val in e.raw.items():
            if tag not in attr_tags:
                attr_tags.append(tag)
            add_val(tag, val)

    fields = sorted(attr_tags) + _DERIVED_FIELDS
    return {
        "ok": True,
        "elements": len(elements),
        "types": sorted(types),
        "layers": sorted(layers),
        "fields": fields,                       # доступно для фильтра/группировки
        "values": {k: sorted(v) for k, v in values.items()},
        "presets": config.get("reports", []),   # пресеты = текущие отчёты
    }


def engine_json(payload: Any, config: dict) -> dict:
    """Точка входа движка для C#-шима.

    Совместимо со старым форматом (payload = список записей блоков -> все пресеты).
    Новый формат — словарь:
      {"blocks":[...], "action":"describe"}                      -> метаданные
      {"blocks":[...], "action":"run", "report":"<имя пресета>"} -> один пресет
      {"blocks":[...], "action":"run", "query":{...}}            -> произвольный запрос
      {"blocks":[...], "action":"run"}                            -> все пресеты
    """
    # старый формат: голый список
    if isinstance(payload, list):
        records = payload
        action, report, query = "run", None, None
    else:
        records = payload.get("blocks", [])
        action = payload.get("action", "run")
        report = payload.get("report")
        query = payload.get("query")

    mapper = ElementMapper(config)
    blocks = [RawBlock.from_record(r) for r in records]
    elements, skipped = mapper.map_all(blocks)
    base = {"ok": True, "blocks_in": len(records),
            "elements": len(elements), "skipped": len(skipped)}

    if action == "describe":
        d = describe(elements, config)
        d.update({"blocks_in": len(records), "skipped": len(skipped)})
        return d

    # «Свой отчёт» (аналог Шаблона отчёта СПДС): источник-фильтр + выражения по
    # столбцам + группировка/сортировка + производные строки. Работает с СЫРЫМИ
    # записями блоков (любое поле/атрибут/слой), без проекции через mapping.yaml.
    if action == "report":
        report_def = payload.get("report") or {}
        base["report"] = atspec_report.run_report(records, report_def)
        return base

    builder = SpecBuilder(config)
    if query:
        # произвольный запрос из диалога
        rc = {
            "name": query.get("name", "запрос"),
            "title": query.get("title", "Ведомость"),
            "include_types": query.get("include_types", []),
            "include_layers": query.get("include_layers", []),
            "filters": query.get("filters", []),
            "group_by": query.get("group_by") or ["type"],
            "measures": query.get("measures") or {"кол_во": "count"},
            "sort_by": query.get("sort_by") or (query.get("group_by") or ["type"]),
        }
        specs = [builder.build(rc, elements)]
    elif report:
        rc = next((r for r in builder.reports if r.get("name") == report), None)
        if rc is None:
            return {"ok": False, "error": f"пресет не найден: {report}", **base}
        specs = [builder.build(rc, elements)]
    else:
        specs = builder.build_all(elements)

    base["reports"] = [spec_to_dict(s) for s in specs]
    return base


# --------------------------------------------------------------------------- #
#  CLI / Pipeline
# --------------------------------------------------------------------------- #
def run(dxf_path: str, config_path: str, out_dir: str,
        xlsx: bool = False, quiet: bool = False) -> List[Spec]:
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    blocks = BlockExtractor(dxf_path).extract()
    mapper = ElementMapper(config)
    elements, skipped = mapper.map_all(blocks)
    specs = SpecBuilder(config).build_all(elements)

    os.makedirs(out_dir, exist_ok=True)
    write_raw_csv(elements, os.path.join(out_dir, "00_raw_elements.csv"))
    for s in specs:
        write_csv(s, os.path.join(out_dir, f"{s.name}.csv"))
    if xlsx:
        write_xlsx(specs, os.path.join(out_dir, "specs.xlsx"))

    if not quiet:
        print(f"Файл: {os.path.basename(dxf_path)}")
        print(f"Вставок всего: {len(blocks)} | элементов: {len(elements)} | "
              f"пропущено (неэлементные слои): {len(skipped)}")
        for s in specs:
            print_spec(s)
        print(f"\nВыгружено в: {out_dir}")
    return specs


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Извлечение и агрегация блоков DXF без СПДС. "
                    "Режимы: файловый (DXF->CSV/XLSX) и гибридный (JSON<->движок).")
    p.add_argument("dxf", nargs="?", help="путь к .dxf (файловый режим и --emit-records)")
    p.add_argument("-c", "--config", required=True, help="YAML-конфиг маппинга/отчётов")
    p.add_argument("-o", "--out", default="out", help="каталог вывода (файловый режим)")
    p.add_argument("--xlsx", action="store_true", help="дополнительно выгрузить specs.xlsx")
    p.add_argument("-q", "--quiet", action="store_true", help="без печати в консоль")
    # --- гибридные режимы ---
    p.add_argument("--emit-records", action="store_true",
                   help="прочитать DXF и вывести в stdout JSON-массив записей блоков "
                        "(формат, который отдаёт C#-плагин из живого чертежа)")
    p.add_argument("--json", action="store_true",
                   help="РЕЖИМ ДВИЖКА: прочитать JSON-записи блоков из stdin (или --in), "
                        "вывести JSON со спецификациями в stdout (или --out-json)")
    p.add_argument("--in", dest="infile", default=None,
                   help="режим движка: читать записи блоков из файла, а не из stdin")
    p.add_argument("--out-json", dest="outjson", default=None,
                   help="режим движка: писать спецификации в файл, а не в stdout")
    a = p.parse_args(argv)
    config = load_config(a.config)

    if a.json:
        # надёжная UTF-8 на стыке с C# (на Windows консоль по умолчанию cp1251)
        for stream in (sys.stdin, sys.stdout):
            try:
                stream.reconfigure(encoding="utf-8")
            except Exception:
                pass
        if a.infile:
            with open(a.infile, "r", encoding="utf-8") as f:
                payload = json.load(f)
        else:
            payload = json.load(sys.stdin)
        # payload: список записей (старый формат) ИЛИ словарь {blocks, action, query/report}
        result = engine_json(payload, config)
        text = json.dumps(result, ensure_ascii=False)
        if a.outjson:
            with open(a.outjson, "w", encoding="utf-8") as f:
                f.write(text)
        else:
            sys.stdout.write(text)
        return

    if a.emit_records:
        if not a.dxf:
            p.error("--emit-records требует путь к .dxf")
        blocks = BlockExtractor(a.dxf).extract()
        json.dump([b.to_record() for b in blocks], sys.stdout, ensure_ascii=False)
        return

    if not a.dxf:
        p.error("укажите .dxf (файловый режим) либо используйте --json (режим движка)")
    run(a.dxf, a.config, a.out, xlsx=a.xlsx, quiet=a.quiet)


if __name__ == "__main__":
    main()
